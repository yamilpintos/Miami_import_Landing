"""
API del panel administrativo, respaldada por la base de datos PROPIA.

Reemplaza al viejo proxy contra Tiendanube. Mantiene los mismos paths y la misma
forma de respuesta (ver serializers.py) para que el frontend siga funcionando.
"""
from __future__ import annotations

import html
import io
import re
import secrets
import unicodedata
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import List, Optional

import requests
from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy import func
from sqlalchemy.orm import Session

from auth import get_current_admin
from core import storage
from core.config import settings
from core.db import get_db
from core.models import (
    Category, Order, Product, ProductImage, Variant, Setting,
)
from core.sanitize import clean_description
from serializers import order_to_tn, product_to_tn

# Todo el panel exige sesión de admin.
router = APIRouter(prefix="/api", tags=["panel"], dependencies=[Depends(get_current_admin)])

STORE_STATIC = Path(__file__).resolve().parent.parent / "web-tienda" / "static"


# --------------------------------------------------------------------------- #
# Helpers
# --------------------------------------------------------------------------- #
def _strip_accents(s: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))


def slugify(text: str, max_len: int = 80) -> str:
    t = _strip_accents(text or "").lower()
    t = re.sub(r"[^a-z0-9]+", "-", t).strip("-")
    return t[:max_len].rstrip("-") or "producto"


def unique_handle(db: Session, base: str) -> str:
    handle = base
    i = 1
    while db.query(Product).filter_by(handle=handle).first():
        i += 1
        handle = f"{base}-{i}"
    return handle


ALLOWED_IMAGE_EXT = {".jpg", ".jpeg", ".png", ".webp", ".gif", ".avif"}
ALLOWED_IMAGE_MIME = {"image/jpeg", "image/png", "image/webp", "image/gif", "image/avif"}
MAX_IMAGE_BYTES = 8 * 1024 * 1024  # 8 MB
# Firmas (magic numbers) para confirmar que el archivo ES una imagen real.
_IMAGE_MAGIC = (b"\xff\xd8\xff", b"\x89PNG\r\n\x1a\n", b"GIF87a", b"GIF89a", b"RIFF")


def validate_image(filename: str, content_type: str, content: bytes) -> str:
    """Valida una imagen subida y devuelve un nombre de archivo SEGURO y aleatorio.
    Rechaza por tamaño, extensión, MIME y firma binaria. Renombrado aleatorio
    evita path traversal y colisiones."""
    if len(content) > MAX_IMAGE_BYTES:
        raise HTTPException(413, "Imagen demasiado grande (máx 8 MB)")
    if len(content) < 64:
        raise HTTPException(400, "Archivo vacío o corrupto")
    ext = Path(filename or "").suffix.lower()
    if ext not in ALLOWED_IMAGE_EXT:
        raise HTTPException(415, f"Extensión no permitida: {ext or '(ninguna)'}")
    if content_type and content_type.lower() not in ALLOWED_IMAGE_MIME:
        raise HTTPException(415, f"Tipo no permitido: {content_type}")
    if not any(content.startswith(sig) for sig in _IMAGE_MAGIC):
        raise HTTPException(415, "El contenido no es una imagen válida")
    return f"{secrets.token_hex(16)}{ext if ext != '.jpeg' else '.jpg'}"


def store_image_bytes(content: bytes, content_type: str, handle: str, safe_name: str) -> tuple[str, str | None]:
    """Guarda la imagen en Supabase Storage (si está configurado) o en disco local.
    Devuelve (src, local_path). En Supabase, src = URL pública y local_path = None."""
    rel = f"products/{handle}/{safe_name}"
    if storage.is_enabled():
        url = storage.upload_bytes(content, rel, content_type or "image/jpeg")
        return url, None
    dest = STORE_STATIC / rel
    dest.parent.mkdir(parents=True, exist_ok=True)
    dest.write_bytes(content)
    return f"/static/{rel}", f"/static/{rel}"


def current_usd_rate(db: Session) -> float:
    s = db.get(Setting, "usd_rate")
    if s and s.value and s.value.get("rate"):
        try:
            return float(s.value["rate"])
        except (TypeError, ValueError):
            pass
    return settings.USD_TO_ARS_RATE


# --------------------------------------------------------------------------- #
# Productos
# --------------------------------------------------------------------------- #
@router.get("/store")
def store_info(db: Session = Depends(get_db)):
    """Datos de la tienda para el frontend del panel.

    El panel viejo tenía la URL de Tiendanube hardcodeada; ahora sale de la
    config para que los links apunten a NUESTRA tienda.
    """
    base = settings.STORE_BASE_URL.rstrip("/")
    return {
        "name": "MIAMI IMPORT",
        "url": base,
        "product_url_base": f"{base}/productos/",
        "usd_rate": current_usd_rate(db),
    }


@router.get("/products")
def list_products(q: Optional[str] = None, page: int = 1, per_page: int = 200,
                  db: Session = Depends(get_db)):
    query = db.query(Product)
    if q:
        like = f"%{q.lower()}%"
        query = query.filter(func.lower(Product.name).like(like) | func.lower(Product.brand).like(like))
    items = query.order_by(Product.id.desc()).offset((page - 1) * per_page).limit(per_page).all()
    return [product_to_tn(p) for p in items]


@router.get("/products/all")
def list_all_products(db: Session = Depends(get_db)):
    return [product_to_tn(p) for p in db.query(Product).order_by(Product.id.desc()).all()]


@router.get("/products/{pid}")
def get_product(pid: int, db: Session = Depends(get_db)):
    p = db.get(Product, pid)
    if not p:
        raise HTTPException(404, "Producto no encontrado")
    return product_to_tn(p)


@router.put("/products/{pid}")
def update_product(pid: int, body: dict, db: Session = Depends(get_db)):
    """Actualiza datos del producto (nombre, marca, descripción, publicado)."""
    p = db.get(Product, pid)
    if not p:
        raise HTTPException(404, "Producto no encontrado")

    if body.get("name") is not None:
        nombre = str(body["name"]).strip()[:500]
        if not nombre:
            raise HTTPException(400, "El nombre no puede quedar vacío")
        p.name = nombre
    if "brand" in body:
        p.brand = str(body.get("brand") or "").strip()[:255] or None
    if "description" in body:
        # Se renderiza con `| safe` en la tienda: sanitizar siempre.
        p.description = clean_description(str(body.get("description") or ""))
    if "published" in body:
        p.published = bool(body["published"])

    db.commit()
    db.refresh(p)
    return product_to_tn(p)


@router.post("/products/{pid}/variants")
def add_variant(pid: int, body: dict, db: Session = Depends(get_db)):
    """Agrega un talle nuevo a un producto existente.

    Se crea con stock 0 por defecto: solo deja la opción disponible, sin tocar
    las variantes que el producto ya tiene cargadas.
    """
    p = db.get(Product, pid)
    if not p:
        raise HTTPException(404, "Producto no encontrado")

    talle = str(body.get("talle") or body.get("value") or "").strip().upper()[:255]
    if not talle:
        raise HTTPException(400, "Falta el talle")
    if any((v.value or "").strip().upper() == talle for v in p.variants):
        raise HTTPException(409, f"El talle {talle} ya existe en este producto")

    try:
        stock = max(0, int(body.get("stock", 0) or 0))
    except (TypeError, ValueError):
        raise HTTPException(400, "Stock inválido")

    # Precio: el del body, o el de la primera variante existente.
    price = None
    if body.get("price") not in (None, ""):
        try:
            price = Decimal(str(body["price"]))
        except (InvalidOperation, TypeError, ValueError):
            raise HTTPException(400, "Precio inválido")
    base = p.variants[0] if p.variants else None
    if price is None:
        price = base.price if base else Decimal("0")

    v = Variant(
        product_id=p.id, value=talle, stock=stock, price=price,
        usd_price=(base.usd_price if base else None),
        compare_at_price=(base.compare_at_price if base else None),
        position=(max((x.position or 0) for x in p.variants) + 1) if p.variants else 1,
        visible=True,
    )
    db.add(v)
    db.commit()
    db.refresh(p)
    return product_to_tn(p)


@router.post("/products/{pid}/images/{image_id}/rotate")
def rotate_product_image(pid: int, image_id: int, body: dict, db: Session = Depends(get_db)):
    """Rota una imagen ya subida y la reemplaza en el mismo lugar.

    A diferencia del panel viejo (que tenía que borrar y resubir a Tiendanube),
    acá se reescribe el archivo en nuestro bucket y se conserva la posición.
    """
    p = db.get(Product, pid)
    img = db.get(ProductImage, image_id)
    if not p or not img or img.product_id != pid:
        raise HTTPException(404, "Imagen no encontrada")
    try:
        degrees = int(body.get("degrees", 0) or 0) % 360
    except (TypeError, ValueError):
        degrees = 0
    if degrees == 0:
        return {"ok": True, "unchanged": True}
    if not storage.is_enabled():
        raise HTTPException(503, "Storage no configurado")

    try:
        from PIL import Image
    except ImportError:
        raise HTTPException(503, "Falta Pillow para rotar imágenes")

    # Descargar la actual, rotar y volver a subir con nombre nuevo.
    try:
        resp = requests.get(img.url, timeout=30)
        resp.raise_for_status()
        src = Image.open(io.BytesIO(resp.content))
        fmt = (src.format or "JPEG").upper()
        if fmt not in ("JPEG", "PNG", "WEBP"):
            fmt = "JPEG"
        # expand=True para que no recorte al rotar 90/270.
        out = src.rotate(-degrees, expand=True)
        if fmt == "JPEG" and out.mode in ("RGBA", "P"):
            out = out.convert("RGB")
        buf = io.BytesIO()
        out.save(buf, format=fmt, quality=90)
        data = buf.getvalue()
    except HTTPException:
        raise
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(502, f"No se pudo procesar la imagen: {exc}")

    ext = {"JPEG": "jpg", "PNG": "png", "WEBP": "webp"}[fmt]
    rel = f"products/{slugify(p.name or 'producto')}/{secrets.token_hex(16)}.{ext}"
    try:
        new_url = storage.upload_bytes(
            data, rel, f"image/{'jpeg' if ext == 'jpg' else ext}")
    except RuntimeError as exc:
        raise HTTPException(502, f"No se pudo guardar la imagen rotada: {exc}")

    old_path = storage.path_from_url(img.url)
    img.url = new_url
    db.commit()
    if old_path:
        try:
            storage.delete_path(old_path)  # recién ahora, con la nueva ya guardada
        except Exception:  # noqa: BLE001
            pass
    return {"ok": True, "src": new_url}


@router.delete("/products/{pid}")
def delete_product(pid: int, db: Session = Depends(get_db)):
    p = db.get(Product, pid)
    if not p:
        raise HTTPException(404, "Producto no encontrado")
    # Borrar también las imágenes del bucket de Supabase (si aplica).
    if storage.is_enabled():
        for img in p.images:
            path = storage.path_from_url(img.src)
            if path:
                storage.delete_path(path)
    db.delete(p)
    db.commit()
    return {"ok": True}


@router.put("/variants/{pid}/{vid}/stock")
def update_variant_stock(pid: int, vid: int, body: dict, db: Session = Depends(get_db)):
    v = db.get(Variant, vid)
    if not v or v.product_id != pid:
        raise HTTPException(404, "Variante no encontrada")
    v.stock = int(body.get("stock", 0))
    db.commit()
    return {"ok": True, "stock": v.stock}


@router.put("/variants/{pid}/{vid}")
def update_variant(pid: int, vid: int, body: dict, db: Session = Depends(get_db)):
    v = db.get(Variant, vid)
    if not v or v.product_id != pid:
        raise HTTPException(404, "Variante no encontrada")
    if "stock" in body:
        v.stock = int(body["stock"])
    if "price" in body:
        v.price = Decimal(str(body["price"]))
    if "sku" in body:
        v.sku = body["sku"]
    if "promotional_price" in body and body["promotional_price"] not in (None, ""):
        v.promotional_price = Decimal(str(body["promotional_price"]))
    db.commit()
    return {"ok": True}


@router.post("/catalogo_entrante/subir")
async def catalogo_entrante_subir(
    filename: str = Form(...),
    name: str = Form(...),
    brand: str = Form(""),
    price: str = Form(...),
    talles: str = Form(""),
    stock_por_talle: int = Form(1),
    publicado: bool = Form(True),
    convertir_a_ars: bool = Form(False),
    rotation: int = Form(0),
    db: Session = Depends(get_db),
):
    """Crea el producto a partir de una foto de la bandeja de entrada.

    Reusa exactamente la misma lógica que el alta manual; lo único distinto es
    de dónde sale la imagen. Al terminar archiva la foto para que salga de la
    lista de pendientes.
    """
    from starlette.datastructures import Headers, UploadFile as StarletteUploadFile

    from catalogo_entrante import _safe_path, archivar_subida

    f = _safe_path(filename)
    if not f.exists():
        raise HTTPException(404, "No existe la imagen")
    data = f.read_bytes()

    # Rotación opcional (las fotos del catálogo a veces vienen giradas).
    if rotation % 360:
        try:
            from PIL import Image
            src = Image.open(io.BytesIO(data))
            fmt = (src.format or "JPEG").upper()
            out = src.rotate(-(rotation % 360), expand=True)
            if fmt == "JPEG" and out.mode in ("RGBA", "P"):
                out = out.convert("RGB")
            buf = io.BytesIO()
            out.save(buf, format=fmt, quality=90)
            data = buf.getvalue()
        except ImportError:
            raise HTTPException(503, "Falta Pillow para rotar la imagen")
        except Exception as exc:  # noqa: BLE001
            raise HTTPException(400, f"No se pudo rotar la imagen: {exc}")

    ctype = "image/png" if f.suffix.lower() == ".png" else (
        "image/webp" if f.suffix.lower() == ".webp" else "image/jpeg")
    upload = StarletteUploadFile(
        file=io.BytesIO(data), filename=f.name,
        headers=Headers({"content-type": ctype}),
    )

    result = await create_product(
        name=name, brand=brand, description="", price=price, talles=talles,
        stock_por_talle=stock_por_talle, publicado=publicado,
        images=[upload], convertir_a_ars=convertir_a_ars, db=db,
    )
    archivar_subida(filename)  # recién ahora: si falló el alta, la foto queda
    return result


@router.post("/products")
async def create_product(
    name: str = Form(...),
    brand: str = Form(...),
    description: str = Form(""),
    price: str = Form(...),
    talles: str = Form(""),
    stock_por_talle: int = Form(1),
    publicado: bool = Form(True),
    images: List[UploadFile] = File([]),
    convertir_a_ars: bool = Form(False),
    db: Session = Depends(get_db),
):
    talles_list = [t.strip().upper() for t in re.split(r"[,;]", talles) if t.strip()]
    base_handle = slugify(f"{brand}-{name}")
    handle = unique_handle(db, base_handle)
    sku_base = re.sub(r"[^A-Z0-9]", "", _strip_accents(f"{brand}{name}").upper())[:30] or "PROD"

    try:
        precio_num = Decimal(str(price))
    except Exception:
        raise HTTPException(400, "Precio inválido")
    rate = Decimal(str(current_usd_rate(db)))
    if convertir_a_ars:
        usd_val = precio_num
        precio_ars = (precio_num * rate).quantize(Decimal("0.01"))
    else:
        precio_ars = precio_num.quantize(Decimal("0.01"))
        usd_val = (precio_ars / rate).quantize(Decimal("0.01")) if rate else None

    # La descripción se renderiza con `| safe` en la tienda, así que el HTML
    # que se guarda tiene que ser SOLO el que armamos acá: los valores del
    # formulario van escapados. Sin esto, un `<script>` guardado desde el panel
    # queda persistente en el dominio donde vive el checkout de Stripe.
    desc = f"<p>{html.escape(name)}</p>"
    if brand:
        desc += f"<p>Marca: {html.escape(brand)}</p>"
    desc += (f"<p>{html.escape(description)}</p>" if description
             else "<p>Producto original importado.</p>")

    prod = Product(name=name, handle=handle, description=desc, brand=brand, published=publicado)
    db.add(prod)
    db.flush()

    talles_iter = talles_list or [None]
    for i, t in enumerate(talles_iter, 1):
        db.add(Variant(
            product_id=prod.id, price=precio_ars, usd_price=usd_val, stock=stock_por_talle,
            sku=f"{sku_base}-{slugify(t).upper()}" if t else sku_base, value=t, position=i,
        ))
    db.flush()

    imgs_ok = 0
    for i, upfile in enumerate(images, 1):
        content = await upfile.read()
        safe_name = validate_image(upfile.filename or "", upfile.content_type or "", content)
        src, local_path = store_image_bytes(content, upfile.content_type or "", handle, safe_name)
        db.add(ProductImage(product_id=prod.id, src=src, local_path=local_path,
                            position=i, alt=name))
        imgs_ok += 1
    db.commit()

    return {
        "ok": True,
        "product_id": prod.id,
        "url": f"{settings.STORE_BASE_URL}/productos/{handle}/",
        "imagenes_subidas": imgs_ok,
        "variantes_creadas": len(talles_iter),
    }


@router.post("/products/{pid}/images")
async def upload_image(pid: int, file: UploadFile = File(...), position: int = Form(99),
                       db: Session = Depends(get_db)):
    p = db.get(Product, pid)
    if not p:
        raise HTTPException(404, "Producto no encontrado")
    content = await file.read()
    safe_name = validate_image(file.filename or "", file.content_type or "", content)
    src, local_path = store_image_bytes(content, file.content_type or "", p.handle, safe_name)
    db.add(ProductImage(product_id=p.id, src=src, local_path=local_path,
                        position=position, alt=p.name))
    db.commit()
    return {"ok": True, "src": src}


# --------------------------------------------------------------------------- #
# Pedidos / estadísticas
# --------------------------------------------------------------------------- #
@router.get("/orders/{oid}")
def get_order(oid: int, db: Session = Depends(get_db)):
    o = db.get(Order, oid)
    if not o:
        raise HTTPException(404, "Pedido no encontrado")
    return order_to_tn(o)


@router.post("/orders/{oid}/status")
def set_order_status(oid: int, body: dict, db: Session = Depends(get_db)):
    from core.models import ORDER_STATUSES
    o = db.get(Order, oid)
    if not o:
        raise HTTPException(404, "Pedido no encontrado")
    new = body.get("status")
    if new not in ORDER_STATUSES:
        raise HTTPException(400, f"Estado inválido. Opciones: {', '.join(ORDER_STATUSES)}")
    o.status = new
    db.commit()
    return {"ok": True, "status": o.status}


@router.post("/orders/{oid}/refund")
def refund_order(oid: int, db: Session = Depends(get_db)):
    """Reembolsa el pago de Stripe asociado a la orden y marca el pedido."""
    o = db.get(Order, oid)
    if not o:
        raise HTTPException(404, "Pedido no encontrado")
    if not settings.STRIPE_SECRET_KEY:
        raise HTTPException(503, "Stripe no está configurado")
    from core.models import Payment
    payment = (
        db.query(Payment).filter(Payment.order_id == o.id,
                                 Payment.stripe_payment_intent_id.isnot(None))
        .order_by(Payment.id.desc()).first()
    )
    if not payment or not payment.stripe_payment_intent_id:
        raise HTTPException(400, "No hay pago de Stripe para reembolsar")
    import stripe
    stripe.api_key = settings.STRIPE_SECRET_KEY
    try:
        refund = stripe.Refund.create(payment_intent=payment.stripe_payment_intent_id)
    except Exception as e:  # noqa: BLE001
        raise HTTPException(502, f"Error de Stripe: {e}")
    payment.status = "refunded"
    o.payment_status = "refunded"
    o.status = "refunded"
    # Reponer stock
    for it in o.items:
        v = db.get(Variant, it.variant_id) if it.variant_id else None
        if v:
            v.stock += it.quantity
    db.commit()
    return {"ok": True, "refund_id": refund.id, "status": "refunded"}


@router.get("/orders")
def list_orders(per_page: int = 50, page: int = 1, status: Optional[str] = None,
                db: Session = Depends(get_db)):
    query = db.query(Order)
    if status:
        query = query.filter(Order.payment_status == status)
    items = query.order_by(Order.created_at.desc()).offset((page - 1) * per_page).limit(per_page).all()
    return [order_to_tn(o) for o in items]


@router.get("/stats")
def stats(db: Session = Depends(get_db)):
    productos = db.query(Product).all()
    total_productos = len(productos)
    total_publicados = sum(1 for p in productos if p.published)
    total_variantes = db.query(Variant).count()
    total_stock = db.query(func.coalesce(func.sum(Variant.stock), 0)).scalar() or 0
    productos_sin_stock = sum(1 for p in productos if p.total_stock == 0)

    pedidos = db.query(Order).all()
    total_pedidos = len(pedidos)
    total_facturado = float(sum((o.total or 0) for o in pedidos))
    pedidos_pagados = sum(1 for o in pedidos if o.payment_status == "paid")
    pedidos_pendientes = sum(1 for o in pedidos if o.payment_status == "pending")

    contar = {}
    for o in pedidos:
        for it in o.items:
            if it.product_id:
                contar[it.product_id] = contar.get(it.product_id, 0) + it.quantity
    top = sorted(contar.items(), key=lambda x: -x[1])[:10]
    name_by_id = {p.id: p.name for p in productos}
    top_named = [{"product_id": pid, "name": name_by_id.get(pid, "?"), "vendidos": q} for pid, q in top]

    stock_bajo = [
        {"id": p.id, "name": p.name, "brand": p.brand, "stock": p.total_stock}
        for p in productos if p.total_stock <= 1
    ][:20]

    return {
        "productos": {
            "total": total_productos, "publicados": total_publicados,
            "sin_stock": productos_sin_stock, "variantes": total_variantes,
            "stock_total": int(total_stock),
        },
        "pedidos": {
            "total": total_pedidos, "pagados": pedidos_pagados, "pendientes": pedidos_pendientes,
            "facturado_total": total_facturado,
            "ticket_promedio": (total_facturado / total_pedidos) if total_pedidos else 0,
        },
        "top_vendidos": top_named,
        "stock_bajo": stock_bajo,
    }


# --------------------------------------------------------------------------- #
# Precios USD (ahora 100% sobre la DB propia)
# --------------------------------------------------------------------------- #
@router.get("/usd_prices")
def usd_prices_get(db: Session = Depends(get_db)):
    prices = {}
    for p in db.query(Product).all():
        v = p.variants[0] if p.variants else None
        if v and v.usd_price is not None:
            prices[str(p.id)] = float(v.usd_price)
    return {"prices": prices, "rate": current_usd_rate(db)}


@router.post("/usd_prices")
def usd_prices_save(body: dict, db: Session = Depends(get_db)):
    prices = body.get("prices") or {}
    saved = 0
    for pid, usd in prices.items():
        try:
            usd_d = Decimal(str(usd))
        except Exception:
            continue
        p = db.get(Product, int(pid))
        if not p:
            continue
        for v in p.variants:
            v.usd_price = usd_d
        saved += 1
    db.commit()
    return {"ok": True, "saved_count": saved}


@router.post("/usd_prices/from_current")
def usd_prices_seed(db: Session = Depends(get_db)):
    rate = Decimal(str(current_usd_rate(db)))
    if rate <= 0:
        raise HTTPException(400, "USD rate inválido")
    count = 0
    for p in db.query(Product).all():
        for v in p.variants:
            if v.price:
                v.usd_price = (v.price / rate).quantize(Decimal("0.01"))
        count += 1
    db.commit()
    return {"ok": True, "count": count, "rate": float(rate)}


@router.post("/usd_prices/sync_to_tiendanube")
def usd_prices_sync(db: Session = Depends(get_db)):
    """Recalcula ARS = USD × rate en TODAS las variantes (ya no toca Tiendanube,
    actualiza nuestra propia base — el nombre se mantiene por compatibilidad del frontend)."""
    rate = Decimal(str(current_usd_rate(db)))
    updated_products = updated_variants = 0
    for p in db.query(Product).all():
        touched = False
        for v in p.variants:
            if v.usd_price is not None:
                v.price = (v.usd_price * rate).quantize(Decimal("0.01"))
                updated_variants += 1
                touched = True
        if touched:
            updated_products += 1
    db.commit()
    return {"ok": True, "updated_products": updated_products,
            "updated_variants": updated_variants, "rate": float(rate)}


# --------------------------------------------------------------------------- #
# Export Excel (desde la DB)
# --------------------------------------------------------------------------- #
@router.get("/export/excel")
def export_excel(db: Session = Depends(get_db)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    productos = db.query(Product).order_by(Product.id.desc()).all()
    pedidos = db.query(Order).order_by(Order.created_at.desc()).limit(1000).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "productos"
    ws.append(["id", "nombre", "marca", "publicado", "url", "stock_total",
               "variantes", "precio_min", "categorías"])
    for p in productos:
        cats = " | ".join(c.name for c in p.categories)
        ws.append([p.id, p.name, p.brand, p.published,
                   f"{settings.STORE_BASE_URL}/productos/{p.handle}/",
                   p.total_stock, len(p.variants),
                   float(p.min_price) if p.min_price else 0, cats])

    wsv = wb.create_sheet("variantes")
    wsv.append(["product_id", "product_name", "variant_id", "sku", "talle", "precio", "stock"])
    for p in productos:
        for v in p.variants:
            wsv.append([p.id, p.name, v.id, v.sku, v.value,
                        float(v.price) if v.price else 0, v.stock])

    wsp = wb.create_sheet("pedidos")
    wsp.append(["id", "numero", "fecha", "estado", "pago", "total", "moneda",
                "cliente", "email", "telefono", "productos"])
    for o in pedidos:
        prods = " | ".join(f"{it.product_name} x{it.quantity}" for it in o.items)
        wsp.append([o.id, o.number, o.created_at.isoformat() if o.created_at else "",
                    o.status, o.payment_status, float(o.total or 0), o.currency,
                    o.contact_name, o.email, o.contact_phone, prods])

    for sh in (ws, wsv, wsp):
        for cell in sh[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="111111", end_color="111111", fill_type="solid")
        sh.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=miami_import_export_{stamp}.xlsx"},
    )
